#!/usr/bin/env python3
"""data/muni-contacts.json を Excel (.xlsx) に整形して書き出す。

scripts/build-muni-contacts.ts が作る CSV/JSON は機械向けなので、
人が眺めて DM の宛先を選ぶための体裁を付ける:
  - シート「優先リスト」(全1,739件) / 「最優先A」(直近1年50件以上) / 「サマリー」
  - 先頭行固定 + オートフィルタ + 列幅調整
  - 段 (A/B/C/D) を色分け、URL はクリック可能に
  - 末尾に送付管理用の空欄 (送付日 / 方法 / 反応メモ)

実行: python3 scripts/export-muni-contacts-xlsx.py
出力: data/muni-contacts.xlsx
"""
import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "muni-contacts.json"
PREF_SRC = ROOT / "data" / "pref-contacts.json"
OUT = ROOT / "data" / "muni-contacts.xlsx"

# (見出し, JSON キー, 列幅, リンクにするか)
COLUMNS = [
    ("優先度", "priority", 7, False),
    ("段", "tier", 5, False),
    ("都道府県", "prefName", 10, False),
    ("市区町村", "cityName", 16, False),
    ("団体コード", "cityCode", 10, False),
    ("出没_直近1年", "sightings365", 12, False),
    ("出没_累計", "sightingsTotal", 10, False),
    ("担当課", "deptName", 34, False),
    ("電話", "tel", 20, False),
    ("FAX", "fax", 16, False),
    ("メール", "email", 30, False),
    ("メール宛先", "emailKind", 10, False),
    ("問い合わせフォーム", "contactFormUrl", 30, True),
    ("郵便番号", "postalCode", 10, False),
    ("住所", "address", 40, False),
    ("到達手段", "reach", 22, False),
    ("郵送宛名", "postalName", 34, False),
    ("公式HP", "homeUrl", 30, True),
    ("クマ情報ページ", "bearUrl", 30, True),
    ("抽出確度", "confidence", 9, False),
    ("抽出日", "extractedAt", 11, False),
    ("備考", "note", 26, False),
]
# 送付管理用の空欄 (こちらで埋めていく列)
TRACK_COLUMNS = [("送付日", 11), ("方法", 12), ("反応メモ", 30)]

# 都道府県シート。市区町村とは列が違う (段でなく県内の状況を持たせる)
PREF_COLUMNS = [
    ("順位", "rank", 6, False),
    ("都道府県", "prefName", 10, False),
    ("生息状況", "bearStatus", 9, False),
    ("県内出没_直近1年", "sightings365", 15, False),
    ("段A市町村", "muniTierA", 10, False),
    ("段B市町村", "muniTierB", 10, False),
    ("市町村数", "muniTotal", 9, False),
    ("担当課", "deptName", 40, False),
    ("電話", "tel", 20, False),
    ("FAX", "fax", 16, False),
    ("メール", "email", 32, False),
    ("問い合わせフォーム", "contactFormUrl", 28, True),
    ("郵便番号", "postalCode", 10, False),
    ("住所", "address", 36, False),
    ("県公式HP", "homeUrl", 28, True),
    ("クマ情報ページ", "bearUrl", 28, True),
    ("抽出確度", "confidence", 9, False),
    ("抽出日", "extractedAt", 11, False),
    ("備考", "note", 24, False),
]

STATUS_FILL = {
    "生息": PatternFill("solid", fgColor="FBE2D5"),
    "希少": PatternFill("solid", fgColor="FDF2D0"),
    "絶滅": PatternFill("solid", fgColor="F2F2F2"),
    "生息なし": PatternFill("solid", fgColor="F2F2F2"),
}

TIER_FILL = {
    "A": PatternFill("solid", fgColor="FBE2D5"),
    "B": PatternFill("solid", fgColor="FDF2D0"),
    "C": PatternFill("solid", fgColor="EAF3E0"),
    "D": PatternFill("solid", fgColor="F2F2F2"),
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0B6BCB", underline="single", size=10)


def write_sheet(ws, rows, columns=None, fills=None, fill_key="tier", fill_col=2, freeze="E2"):
    columns = columns or COLUMNS
    fills = TIER_FILL if fills is None else fills
    headers = [c[0] for c in columns] + [c[0] for c in TRACK_COLUMNS]
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for r in rows:
        ws.append([r.get(key, "") for _, key, _, _ in columns] + [""] * len(TRACK_COLUMNS))
        row_idx = ws.max_row
        fill = fills.get(r.get(fill_key, ""))
        for col_idx, (_, key, _, is_link) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("deptName", "address", "note"))
            if is_link and r.get(key):
                cell.hyperlink = r[key]
                cell.font = LINK_FONT
        # 1 列だけ色を付ける (行全体を塗ると読みづらい)
        if fill:
            ws.cell(row=row_idx, column=fill_col).fill = fill

    for col_idx, (_, _, width, _) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for offset, (_, width) in enumerate(TRACK_COLUMNS):
        ws.column_dimensions[get_column_letter(len(columns) + 1 + offset)].width = width

    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"


def write_summary(ws, rows, pref_rows=None):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 52

    def line(a, b="", c="", bold=False):
        ws.append([a, b, c])
        if bold:
            for i in range(1, 4):
                ws.cell(row=ws.max_row, column=i).font = Font(bold=True)

    def has_contact(r):
        return bool(r["tel"] or r["email"] or r["contactFormUrl"])

    line("くまウォッチ 自治体連絡先リスト", "", "", bold=True)
    line("生成日", rows and rows[0].get("extractedAt", ""), "scripts/build-muni-contacts.ts")
    line("")
    line("全体", "", "", bold=True)
    line("自治体数", len(rows), "政令指定都市の区は市に集約")
    line("公式HPあり", sum(1 for r in rows if r["homeUrl"]))
    line("クマ情報ページあり", sum(1 for r in rows if r["bearUrl"]))
    line("連絡先あり", sum(1 for r in rows if has_contact(r)), "電話・メール・フォームのいずれか")
    line("電話", sum(1 for r in rows if r["tel"]))
    line("問い合わせフォーム", sum(1 for r in rows if r["contactFormUrl"]))
    line("メールアドレス", sum(1 for r in rows if r["email"]), "自治体はフォームに統一済みで公開が少ない")
    line("うち担当課に直接届く", sum(1 for r in rows if r.get("emailKind") == "担当課"), "info@/koho@ は総合窓口どまり")
    line("住所（郵送用）", sum(1 for r in rows if r["address"]))
    line("担当課まで特定", sum(1 for r in rows if r["deptName"]))
    line("")
    line("優先度の段（直近1年の出没件数）", "", "", bold=True)
    labels = {"A": "50件以上", "B": "10〜49件", "C": "1〜9件", "D": "0件"}
    for t in ("A", "B", "C", "D"):
        g = [r for r in rows if r["tier"] == t]
        line(f"段 {t}（{labels[t]}）", len(g), f"連絡先あり {sum(1 for r in g if has_contact(r))} 件")
    line("")
    line("都道府県（最初に当たる先）", "", "", bold=True)
    if pref_rows:
        bear = [r for r in pref_rows if r["bearStatus"] in ("生息", "希少")]
        line("クマ生息県", len(bear), "ここが本命。県から市町村へ紹介が回る")
        line("うち担当課まで特定", sum(1 for r in bear if r["deptName"]))
        line("うちメールあり", sum(1 for r in bear if r["email"]))
        line("47都道府県 連絡先あり", sum(1 for r in pref_rows if r["tel"] or r["email"] or r["contactFormUrl"]))
    else:
        line("未生成", "", "npm run build:pref-contacts で作成")
    line("")
    line("出典・注意", "", "", bold=True)
    line("出没件数", "", "public/data/sightings.json（自社集計・直近365日）")
    line("連絡先", "", "各自治体公式サイトの記載を機械抽出。送付前に要確認")
    line("抽出確度 high", "", "クマ・鳥獣の担当部署が明示されていた")
    line("抽出確度 low", "", "代表窓口どまり。担当課は電話で確認が要る")
    line("問い合わせフォーム", "", "自動送信はしていない。人が開いて送る前提")


def main():
    data = json.loads(SRC.read_text(encoding="utf-8"))
    rows = data["rows"]

    wb = Workbook()
    ws_pref = wb.active
    ws_pref.title = "都道府県"
    pref_rows = []
    if PREF_SRC.exists():
        pref_rows = json.loads(PREF_SRC.read_text(encoding="utf-8"))["rows"]
        write_sheet(
            ws_pref, pref_rows, columns=PREF_COLUMNS, fills=STATUS_FILL,
            fill_key="bearStatus", fill_col=3, freeze="C2",
        )
    else:
        ws_pref["A1"] = "data/pref-contacts.json がありません (npm run build:pref-contacts)"

    write_sheet(wb.create_sheet("優先リスト"), rows)

    tier_a = [r for r in rows if r["tier"] == "A"]
    write_sheet(wb.create_sheet("最優先A"), tier_a)

    write_summary(wb.create_sheet("サマリー"), rows, pref_rows)

    wb.save(OUT)
    print(f"[xlsx] {OUT}")
    print(f"  都道府県 {len(pref_rows)} 行 / 優先リスト {len(rows)} 行 / 最優先A {len(tier_a)} 行")


if __name__ == "__main__":
    main()
